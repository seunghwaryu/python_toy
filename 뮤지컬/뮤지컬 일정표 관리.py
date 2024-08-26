from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import pyautogui

# 엑셀 파일이 저장될 위치
file_path = 'D:/Me/뮤지컬 스케줄/%s.xlsx'

# 인터파크에서 공연 상품페이지로 들어가는 함수
def visitInterpark(name):
    driver = webdriver.Chrome()
    driver.get('https://tickets.interpark.com/') # 인터파크 티켓 접속
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    
    # 검색창 찾기
    ele = driver.find_element(By.CSS_SELECTOR,'#__next > div > header > div.header_wrap__abfca > div.header_navi__b5313 > div > div.header_biSearch__0c77a > div.search_wrap__370af > div > input[type=text]') 
    ele.send_keys(name) # 검색어 입력
    # 검색 버튼 클릭
    pyautogui.press('enter')
    #driver.find_element(By.CSS_SELECTOR,'#__next > div > header > div.GDabBtDh01BTKGQc > div.xKkbLESYEoke3BjV > div > div.eFRFqYToyYg7t4IF > div.FI4lWaRTIr0rIskw > div > button').click() 
    # 검색해서 나오는 첫번째 공연 클릭
    driver.find_element(By.CSS_SELECTOR,'#contents > div > div > div.result-ticket_wrapper__.result-ticket_wrapper__H41_U > div.result-ticket_listWrapper__xcEo3.InfiniteList_list__3c511.InfiniteList_column-desktop-4__c6aac.InfiniteList_column-mobile-1__853f7.InfiniteList_ticket-list__dfe68 > a').click()
    # driver.find_element(By.XPATH,'//*[@id="__next"]/div/main/div/div/div[1]/div[2]/a[1]').click()

    # 현재 창의 핸들을 저장
    main_window = driver.current_window_handle
    driver.close() # 검색 페이지 닫기

    # 검색 페이지 창에서 공연 상품페이지 창으로 전환
    for handle in driver.window_handles:
        if handle != main_window:
            driver.switch_to.window(handle)
            break
    
    # 여기서부터 공연 상품페이지에서 작업 이루어짐
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    try:
        driver.find_element(By.CSS_SELECTOR,'#popup-prdGuide > div > div.popupFooter > button').click() # 팝업 창 있으면 닫기
    except:
        pass
    
    return driver
        
    
# 인터파크에서 전체 일정표 가져오는 함수
def getScheduleFromWeb(name):
    driver = visitInterpark(name)
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    driver.find_element(By.CSS_SELECTOR,'#productMainBody > nav > ul > li:nth-child(2)').click() # 캐스트정보 클릭
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    tbody = driver.find_element(By.CSS_SELECTOR,'#productMainBody > div > div > div.castingDetailResult > table > tbody') # 일정표 전체 얻기

    rows = tbody.find_elements(By.TAG_NAME,'tr') # 표의 열 얻기
    columns = [r.text for r in rows[0].find_elements(By.TAG_NAME,'th')] # 표의 열 이름 저장
    schedule_info = []

    # 일정표의 정보 가져와서 2차원 배열로 저장
    for r in rows[1::]: 
        values = [e.text for e in r.find_elements(By.TAG_NAME,'td')]
        schedule_info.append(values)
    driver.close()

    df = pd.DataFrame(data=schedule_info,columns=columns) # 엑셀로 저장할 수 있게 dataframe으로 만들어주기
    
    return df

# 시트 디자인 설정 함수
def setExcelStyle(ws, need_width):
    default_width = ws.sheet_format.defaultColWidth # 기본 열의 폭 구하기
    if not default_width:
        default_width = 8
        
    cell_color = 'EBF1DE' # 셀 색깔 
    column_color = '9BBB59' # column 색깔
    set_cell_fill = PatternFill(start_color=cell_color, end_color=cell_color,fill_type='solid')
    set_column_fill = PatternFill(start_color=column_color, end_color=column_color,fill_type='solid')
    box = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')) # 테두리 지정
    
    if(need_width > default_width):
        for i in range(2,ws.max_column):
            c = chr(65+i)
            ws.column_dimensions[c].width = need_width
    
    # 모든 셀에 가운데 정렬 및 테두리 설정
    for r in range(ws.max_row):
        for c in range(ws.max_column):
            cell = ws.cell(row=r+1,column=c+1)
            cell.alignment = Alignment(horizontal='center')
            cell.border = box
    
    # column 행에 색 채우기 및 볼드체 설정
    for c in range(ws.max_column):
        cell = ws.cell(row=1,column=c+1)
        cell.font = Font(bold=True)
        cell.fill = set_column_fill
    
    # 짝수번 행에 색 채우기    
    for r in range(1, ws.max_row, 2):
        for c in range(ws.max_column):
            cell = ws.cell(row=r+1,column=c+1)
            cell.fill = set_cell_fill
    
    return ws

# 엑셀파일 불러오는 함수
def loadExcelfile(file_name):
    # file_name과 동일한 이름의 파일이 있으면 그 파일을 불러와 데이터프레임으로 변환한 것을 반환
    # 파일이 없으면 빈 데이터프레임을 반환한다.
    # 메인함수에서의 활용을 위해 pandas로 불러온다.
    try:
        df = pd.read_excel(file_path %file_name)
    except:
        df = pd.DataFrame()
    return df

# 일정표를 엑셀파일로 저장하는 함수
def saveScheduleInExcel(schedule_df, file_name, sheet_name):
    # sheet_name이 전체스케줄이라는 것은 현재 파일이 존재하지 않는 것이다.
    if sheet_name == '전체스케줄':
        wb = Workbook() 
        wb.remove(wb['Sheet'])
    else:
        wb = load_workbook(file_path %file_name)
        
    wb.create_sheet(sheet_name) # 새로운 시트 생성
    ws = wb[sheet_name] # 앞으로 다루는 시트를 앞서 생성한 시트로 설정
    
    need_width = max([len(c) for c in schedule_df.columns])*2 # columns의 최대글자수에 폭을 맞추기 위한 너비 
    
    # 시트에 일정표 데이터 추가    
    for r in dataframe_to_rows(schedule_df, index=False, header=True):
        ws.append(r)
        
    ws = setExcelStyle(ws, need_width) # 시트 스타일 설정
    wb.save(file_path %file_name) # 엑셀파일로 저장하기

# 원하는 배역:배우 입력받는 함수
def inputCast(cast_list):
    cast_dict = dict()
    print('\n원하는 배우와 그 배우의 배역명을 입력해주세요. 입력 종료를 원하면 0을 입력해주세요. ex) 장발장 민우혁')
    print("배역명 목록: ",end='')
    print(*cast_list,sep=', ')
    print('-' * 100)
    
    while True:
        cast = input('입력: ').split()
        if(cast[0] == '0'):
            break
        cast_name = ' '.join(cast[:-1:])
        cast_dict[cast_name] = cast[-1]
    
    return cast_dict

# 일정표 필터링 하는 함수
def fillterSchdeuleByActor(cast_dict, file_name):
    df = pd.read_excel(file_path %file_name) # 파일 불러오기
    # 필터링
    for name, actor in cast_dict.items(): 
        df = df[df[name] == actor] # 필터링한 데이터프레임을 그대로 다시 저장
    
    return df

# 날짜 비교 함수
def compareDate(str_date):
    today = datetime.now()
    last_date = datetime.strptime(str_date,'%m/%d') # 문자열을 datetime 형식으로 바꿔주기
    
    # last_date는 년도가 설정되있지 않기에 현재의 월과 last_date의 월을 비교하기 위해 월의 값을 추출
    # last_date의 년도 설정을 위해 현재 년도 추출
    now_year = int(today.strftime('%Y'))
    now_month = int(today.strftime('%m'))
    date_month = int(last_date.strftime('%m'))
    
    # 뮤지컬은 보통 3개월 정도 하기 때문에 현재 10월이고 1월인 일정이 있다면 그 1월은 내년 1월을 뜻하기에 다음과 같이 년도 설정
    if now_month > 10 and date_month < 3:
        last_date = last_date.replace(year = now_year + 1)
    else:
        last_date = last_date.replace(year = now_year)
    
    # 메인에서 활용하기 위해 True와 False를 반환
    if today > last_date:
        return True
    else:
        return False
    
# 인터파크에서 가격 정보 가져오는 함수
def getPriceFromInterpark(name):
    driver = visitInterpark(name)
    # 전체가격버튼 클릭
    driver.find_element(By.CSS_SELECTOR,'#container > div.contents > div.productWrapper > div.productMain > div.productMainTop > div > div.summaryBody > ul > li.infoItem.infoPrice > div > ul > li.infoPriceItem.is-largePrice > a').click()
    # 가격정보 전체 얻기
    tbody = driver.find_element(By.CSS_SELECTOR,'#popup-info-price > div > div.popupBody > div > div > table > tbody')
    rows = tbody.find_elements(By.TAG_NAME,'tr') # 표의 열 얻기
   
    price_list = []
    # 가격정보 가져와서 2차원 배열로 저장
    for r in rows: 
            values = [e.text for e in r.find_elements(By.TAG_NAME,'td')]
            price_list.append(values)
    driver.close()
    df = pd.DataFrame(data = price_list)
    
    # 할인명에 따른 가격 정리
    seat = ""
    seat_list = []
    price_info = dict() # {할인명1: {등급:가격}, 할인명2: {등급:가격},....}
    for i in range(len(df)):
        n = df[0][i].split('\n')[0] 
        p = df[1][i]
        op = df[2][i]
        
        if op: # 정가일 경우
            seat = n
            seat_list.append(seat)
            if p in price_info:
                price_info[p][seat] = op
            else:
                price_info[p] = {seat:op}
        else: # 할인가 일 경우
            if n in price_info:
                price_info[n][seat] = p
            else:
                price_info[n] = {seat:p}  
                
    discount_list = list(price_info.keys())
    seat_and_price = list(price_info.values())
    price_info = {'할인명':discount_list} # {할인명: 할인명리스트, vip석: 가격리스트,...}
    
    # 각 좌석 등급에 대한 빈 리스트 초기화
    for s in seat_list:
        price_info[s] = []
    
    # 할인명에 따른 좌석등급별 금액으로 정리
    for arr in seat_and_price:
        for s in seat_list:
            if s in arr:
                price_info[s].append(arr[s])
            else:
                price_info[s].append('0원')
                
    df = pd.DataFrame(price_info)
    df = df.drop_duplicates(ignore_index = True) # 할인명 중복 제거
    
    return df

# KT 멤버십에서 가격 정보 가져오는 함수
def getPriceFromKT(name):
    df = pd.DataFrame()
    driver = webdriver.Chrome()
    driver.get('https://membership.kt.com/culture/show/BookingInfo.do') # kt 멤버십 공연 페이지 접속
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    
    # 프레임 이동
    content = driver.find_element(By.TAG_NAME,"iframe") 
    driver.switch_to.frame(content)
    
    # 검색어 입력
    driver.find_element(By.XPATH,'/html/body/div/div/div/div[2]/div[1]/div/fieldset/input').send_keys(name) 
    # 검색 버튼 클릭
    driver.find_element(By.CSS_SELECTOR,'#sub-culture > div.form-thumbnail-box > div.category > div > fieldset > button').click() 
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    
    # 검색 결과가 있으면 가격 정보 가져오고 아니면 빈 dataframe 반환
    try:
        driver.find_element(By.CSS_SELECTOR,'#sub-culture > div.form-thumbnail-box > div.thumbnail.figcaption.musical-figcaption > ul > li > a').click()
    except:
        driver.close()
        return df
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    
    thead = driver.find_element(By.CSS_SELECTOR,'#sub-culture > table > thead') # 표의 열 이름 가져오기
    columns = [c.text for c in thead.find_elements(By.TAG_NAME,'th')] # 표의 열 이름 저장
    tbody = driver.find_element(By.CSS_SELECTOR,'#sub-culture > table > tbody') # 가격정보 얻기
    rows = tbody.find_elements(By.TAG_NAME,'tr') # 표의 열 얻기
    
    price_list =[]
    for r in rows: 
        values = [e.text for e in r.find_elements(By.TAG_NAME,'td')]
        price_list.append(values)
        
    driver.close()
    df = pd.DataFrame(data = price_list, columns = columns)

    # 할인명을 중복없이 추출하여 저장
    price_info = {'할인명': list(set(df['가격등급']))}

    # 각 좌석 등급에 대한 빈 리스트 초기화
    for i in range(len(df)):
        price_info[df['좌석등급'][i]] = []

    # 할인된 가격을 각 좌석 등급에 맞게 추가
    for discount_name in price_info['할인명']:
        for j in range(len(df)):
            # 현재 가격 등급이 할인명과 같다면 할인가를 해당 좌석 등급에 추가
            if df['가격등급'][j] == discount_name:
                price_info[df['좌석등급'][j]].append(df['할인가'][j])
            else:
                # 아니면 0을 추가
                price_info[df['좌석등급'][j]].append('0원')

    df = pd.DataFrame(price_info)
    return df

def getPriceFromWemake(name):
    df = pd.DataFrame()
    driver = webdriver.Chrome()
    driver.get('https://ticket.wemakeprice.com/') # 위메프 티켓 접속
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    
    # 검색어 입력
    driver.find_element(By.CSS_SELECTOR,'#srh-input').send_keys(name) 
    # 겁색 버튼 클릭
    driver.find_element(By.CSS_SELECTOR,'#header > div.u-global-width > div.srh-area > form > fieldset > button').click()
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    
    # 검색 결과가 있으면 가격 정보 가져오고 아니면 빈 dataframe 반환
    try:
        driver.find_element(By.CSS_SELECTOR,'#prodList > ul > li > a > div.cont-wrap').click()
    except:
        driver.close()
        return df
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    
    # 팝업창 있으면 닫기
    try:
        driver.find_element(By.CSS_SELECTOR,'#btn_layerpopup_close').click()
    except:
        pass
    
    # 가격정보 가져오기
    tbody = driver.find_element(By.CSS_SELECTOR,'#basicPrice')
    rows = tbody.find_elements(By.CLASS_NAME,'price-n') 
    price_list = [r.text.split('  ') for r in rows]
    driver.close()
    
    price_info = {'할인명':['위메프 할인']}
    price_info.update({i[0]:[''.join(i[1].split())] for i in price_list})
    
    df = pd.DataFrame(price_info)
    return df

# 가격정보 가져오는 함수
def getPrice(name):
    df_list = [] # dataframe 합치기위한 리스트
    
    # 각 사이트에서 가격 정보 가져오기
    temp_df = getPriceFromInterpark(name)
    temp_df['예매처'] = ['인터파크' for _ in range(len(temp_df))]
    df_list.append(temp_df)

    temp_df = getPriceFromKT(name)
    if not temp_df.empty:
        temp_df['예매처'] = ['KT 멤버십' for _ in range(len(temp_df))]
        df_list.append(temp_df)
    
    temp_df = getPriceFromWemake(name)
    if not temp_df.empty:
        temp_df['예매처'] = ['위메프' for _ in range(len(temp_df))]
        df_list.append(temp_df)

    df = pd.concat(df_list, ignore_index=True) # 가격정보 하나로 합치기
    df.fillna('0원', inplace=True) 
    
    # 제일 높은 좌석등급을 기준으로 가격 정렬하기 
    df = df.sort_values(by=df.columns[1], key=lambda x: x.str.replace(',', '').str.replace('원', '').astype(int), ascending=False)
    return df

# 가격정보를 엑셀로 저장하는 함수
def savePriceInExcel(name):
    file_name = name+'_가격정보'
    
    wb = Workbook()
    ws = wb.active
    ws.title = '가격정보'  
    
    price_df = getPrice(name)
    
    # 시트에 가격 데이터 추가    
    for r in dataframe_to_rows(price_df, index=False, header=True):
        ws.append(r)
        
    ws = setExcelStyle(ws, 12) # 시트 스타일 설정
    wb.save(file_path %file_name) # 엑셀파일로 저장하기
    

# main
input_name = input('정보를 검색할 공연명을 입력해주세요: ')
choice = int(input('1.할인정보 2.일정표: '))

if(choice == 1):
    answer = "Y" 
    # 파일이 이미 있다면
    if not loadExcelfile(input_name+'_가격정보').empty:
        answer = input('가격정보 파일이 이미 존재합니다. 갱신을 원하면 Y을 입력해주세요: ' )
    if answer.upper() == 'Y':    
        savePriceInExcel(input_name)
    print('가격정보의 0원은 해당하는 할인이 없는 것을 의미합니다.')
elif(choice == 2):
    schedule_df = loadExcelfile(input_name)
    
    # 입력받은 공연명과 이름이 동이한 엑셀파일이 존재 하지않다면
    if schedule_df.empty:
        schedule_df = getScheduleFromWeb(input_name)
        saveScheduleInExcel(schedule_df,input_name,'전체스케줄')
    else:
        answer = input('일정표 파일이 이미 존재합니다. 갱신을 원하면 Y을 입력해주세요: ' )
        if(answer.upper() == 'Y'):
            schedule_df = getScheduleFromWeb(input_name)
            saveScheduleInExcel(schedule_df,input_name,'전체스케줄')
        else:
            last_date = schedule_df.iloc[-1][0] # 일정표의 마지막 일정의 날짜 가져오기
            last_date = last_date.split('(')[0] # 요일 정보 삭제
            
            # 현재 가지고 있는 일정표의 마지막 날짜가 현재보다 과거라면
            if compareDate(last_date):
                print("일정표의 정보가 현재 정보가 아님으로 일정표를 다시 가져옵니다.")
                schedule_df = getScheduleFromWeb(input_name)
                saveScheduleInExcel(schedule_df,input_name,'전체스케줄')
    
    # 캐스트 입력받기
    cast_list = list(schedule_df.columns[2:])
    cast_dict = inputCast(cast_list)
    
    # 케스트 입력을 받았다면
    if(len(cast_dict) != 0):
        cast_dict = dict(sorted(cast_dict.items(),key= lambda item:item[1])) # 배우 이름 순으로 정렬
        
        # 필터링하기
        filtered_schedule_df = fillterSchdeuleByActor(cast_dict, input_name)
        
        # 필터링한 일정표 엑셀로 저장
        sheet_name = ','.join(list(cast_dict.values())) # 시트 이름 배우들 이름으로 설정
        saveScheduleInExcel(filtered_schedule_df,input_name,sheet_name)