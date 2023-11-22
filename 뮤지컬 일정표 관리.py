from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 인터파크에서 전체 일정표 가져오는 함수
def getScheduleFromWeb(name):
    
    driver = webdriver.Chrome()
    driver.get('https://tickets.interpark.com/') # 인터파크 티켓 접속
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    
    # 검색창 찾기
    ele = driver.find_element(By.CSS_SELECTOR,'#__next > div > header > div > div._navi_p92f5_16 > div > div._biSearch_p92f5_76 > div._wrap_1iig7_1 > div > input[type=text]') 
    ele.send_keys(name) # 검색어 입력
    # 검색 버튼 클릭
    driver.find_element(By.CSS_SELECTOR,'#__next > div > header > div > div._navi_p92f5_16._autoComplete_p92f5_38 > div > div._biSearch_p92f5_76 > div._wrap_1iig7_1 > div._searchInput_1iig7_16._active_1iig7_33 > button._searchBtn_1iig7_101').click() 
    # 검색해서 나오는 첫번째 공연 클릭
    driver.find_element(By.CSS_SELECTOR,'#__next > div > main > div > div > div.result-ticket_wrapper__H41_U > div.result-ticket_listWrapper__xcEo3 > div:nth-child(1)').click()

    # 현재 창의 핸들을 저장
    main_window = driver.current_window_handle
    driver.close() # 검색 페이지 닫기

    # 검색 페이지 창에서 공연 상품페이지 창으로 전환
    for handle in driver.window_handles:
        if handle != main_window:
            driver.switch_to.window(handle)
            break
    
    # 여기서부터 공연 상품페이지에서 작업 이루어짐
    
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    try:
        driver.find_element(By.CSS_SELECTOR,'#popup-prdGuide > div > div.popupFooter > button').click() # 팝업 창 있으면 닫기
    except:
        pass

    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    driver.find_element(By.CSS_SELECTOR,'#productMainBody > nav > div > div > ul > li:nth-child(2) > a').click() # 캐스트정보 클릭
    driver.implicitly_wait(time_to_wait=5) # 로드될 때까지 대기
    tbody = driver.find_element(By.CSS_SELECTOR,'#productMainBody > div > div > div.castingDetailResult > table > tbody') # 일정표 전체 얻기

    rows = tbody.find_elements(By.TAG_NAME,'tr') # 표의 열 얻기
    columns = [r.text for r in rows[0].find_elements(By.TAG_NAME,'th')] # 표의 열 이름 저장
    schedule_info = []

    # 일정표의 정보 가져와서 2차원 배열로 저장
    for r in rows[1::]: 
        values = [e.text for e in r.find_elements(By.TAG_NAME,'td')]
        schedule_info.append(values)
    driver.close()

    df = pd.DataFrame(data=schedule_info,columns=columns) # 엑셀로 저장할 수 있게 dataframe으로 만들어주기
    
    return df

# 시트 디자인 설정 함수
def setExcelStyle(ws):
    cell_color = 'EBF1DE' # 셀 색깔 
    column_color = '9BBB59' # column 색깔
    set_cell_fill = PatternFill(start_color=cell_color, end_color=cell_color,fill_type='solid')
    set_column_fill = PatternFill(start_color=column_color, end_color=column_color,fill_type='solid')
    box = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')) # 테두리 지정
    
    # 모든 셀에 가운데 정렬 및 테두리 설정
    for r in range(ws.max_row):
        for c in range(ws.max_column):
            cell = ws.cell(row=r+1,column=c+1)
            cell.alignment = Alignment(horizontal='center')
            cell.border = box
    
    # column 행에 색 채우기 및 볼드체 설정
    for c in range(ws.max_column):
        cell = ws.cell(row=1,column=c+1)
        cell.font = Font(bold=True)
        cell.fill = set_column_fill
    
    # 짝수번 행에 색 채우기    
    for r in range(1, ws.max_row, 2):
        for c in range(ws.max_column):
            cell = ws.cell(row=r+1,column=c+1)
            cell.fill = set_cell_fill
                
# 일정표를 엑셀파일로 저장하는 함수
def saveScheduleInExcel(schedule,file_name):
    wb = Workbook()
    ws = wb.active
    
    for r in dataframe_to_rows(schedule, index=False, header=True):
        ws.append(r)
    setExcelStyle(ws)
    wb.save('./%s.xlsx' %file_name)

# 메인 함수
input_name = input('정보를 검색할 공연명을 입력해주세요: ')
schedule = getScheduleFromWeb(input_name)
saveScheduleInExcel(schedule,input_name)
